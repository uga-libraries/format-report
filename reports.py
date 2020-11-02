"""EXPLANATION and prior to running instructions. Explain what the format and usage reports are."""

# Usage: python /path/reports.py ????
# TODO: before delete any previous scripts, read through one more time for ideas for future work.

# TESTING: does this get simpler if using the merged csv organized by aip-format instead of by format? Can I do any
# additional types of analysis?

import csv
import datetime
import openpyxl
import os
import pandas as pd
import sys


def archive_overview():
    """Gets TBs, AIPs, Collections, and Files per group and total for ARCHive using the usage and formats reports.
    Returns the information in a dataframe. """

    def size_and_aips_count():
        """Gets the size in TB and AIP count for each group from the usage report and calculates the total size and AIP
        count for all groups combined. Returns a dictionary with the group code plus total as the keys and lists with
        the group code, size, and number of AIPs as the values. """

        # Group Names maps the human-friendly version of group names from the usage report to the ARCHive group
        # code which is used in the formats report and in ARCHive metadata generally.
        group_names = {'Brown Media Archives': 'bmac', 'Digital Library of Georgia': 'dlg',
                       'DLG & Hargrett': 'dlg-hargrett', 'DLG & Map and Government Information Library': 'dlg-magil',
                       'Hargrett': 'hargrett', 'Russell': 'russell'}

        # Makes a dictionary for storing data for each group that will later be saved to the summary CSV.
        group_data = {}

        # Gets the data from the usage report. Reading this csv instead of making it a dataframe immediately because
        # of all the value updates and the inconsistent values for the different rows.
        with open(usage_report, 'r') as usage:
            usage_read = csv.reader(usage)

            # Skips the header row.
            next(usage_read)

            # Gets data from each row. A row can have data on a group, an individual user, or be blank.
            for row in usage_read:

                # Skips empty rows. Blank rows are used for formatting the usage_report report to be easier to read.
                # Have to do this before the next step or get an IndexError when checking the group name.
                if not row:
                    continue

                # Processes data from each group. There is only one row per group in the report.
                # row[0] is group, row[1] is AIP count, and row[2] is size with the unit of measurement.
                if row[0] in group_names:

                    # Group code is fine as is.
                    group_code = group_names[row[0]]

                    # Changes AIP count from a string to an integer so the total across all groups can be calculated.
                    aip_count = int(row[1])

                    # Separates the size number from the unit of measurement by splitting the data at the space and
                    # converts the size to TB. Example: 100 GB becomes 0.1. All sizes are converted from a string to a
                    # float (decimal numbers) to do the necessary math and to round the result. If it encounters a unit
                    # of measurement that wasn't anticipated, the script prints a warning message.
                    size, unit = row[2].split()
                    if unit == 'Bytes':
                        size = float(size) / 1000000000000
                    elif unit == 'KB':
                        size = float(size) / 1000000000
                    elif unit == 'MB':
                        size = float(size) / 1000000
                    elif unit == 'GB':
                        size = float(size) / 1000
                    elif unit == 'TB':
                        size = float(size)
                    else:
                        size = 0
                        print("WARNING! Unexpected unit type:", unit)

                    # Rounds the size in TB to three decimal places so the number is easier to read.
                    size = round(size, 3)

                    # Adds the results for this group to the dictionary.
                    group_data[group_code] = ([size, aip_count])

        # Makes the dictionary into a dataframe so it can be combined with the collection and file counts.
        group_dataframe = pd.DataFrame.from_dict(group_data, orient='index', columns=['Size', 'AIPs'])

        # Returns the dataframe. Row index is the group_code and columns are the size in TBs and AIPs count.
        return group_dataframe

    # Gets the size (TB) and number of AIPs per group from the usage report.
    size_and_aips_by_group = size_and_aips_count()

    # Gets the number of collections per group from the formats by AIP report.
    # Only counts collections with AIPs, which may result in a difference between this count and ARCHive's count.
    # TODO: add error handling in case AIPs without collections calculated remain in the formats report?
    collections_by_group = df_aip.groupby('Group')['Collection'].nunique()

    # Calculates the number of dlg collections starting 'guan_'. Those should be dlg-hargrett instead.
    # Updates the values in the dataframe to correct for the error.
    dlg = df_aip[df_aip.Group.eq('dlg')]
    guan = pd.Series(dlg['Collection'].unique()).str.startswith('guan_').sum()
    collections_by_group['dlg'] -= guan
    collections_by_group['dlg-hargrett'] += guan

    # Gets the number of files per group from the other formats report.
    # These numbers are inflated by files with more than one format.
    files_by_group = df.groupby('Group')['File_Count'].sum()

    # Combines the dataframes into a single dataframe.
    group_frames = [size_and_aips_by_group, collections_by_group, files_by_group]
    group_combined = pd.concat(group_frames, axis=1)

    # Renames the columns. Only AIPs stays the same.
    group_combined = group_combined.rename(columns={"Size": "Size (TB)", "Collection": "Collections",
                                                    "File_Count": "Files (inflated)"})

    # For Collections and File_Count, replace cells without values (groups that have no collection or files) with 0
    # and returns them to being integers. These counts are initially floats (decimal numbers) because of the blank
    group_combined = group_combined.fillna(0)

    # Adds the column totals to the format type dataframes.
    group_combined.loc['total'] = [group_combined['Size (TB)'].sum(), group_combined['AIPs'].sum(),
                                   group_combined['Collections'].sum(), group_combined['Files (inflated)'].sum()]

    # Makes all rows except size integers, since counts must be whole numbers.
    group_combined['AIPs'] = group_combined['AIPs'].astype(int)
    group_combined['Collections'] = group_combined['Collections'].astype(int)
    group_combined['Files (inflated)'] = group_combined['Files (inflated)'].astype(int)

    # Returns the information in a dataframe
    return group_combined


# Makes the report folder (script argument) the current directory. Displays an error message and quits the script if
# the argument is missing or not a valid directory.
try:
    report_folder = sys.argv[1]
    os.chdir(report_folder)
except (IndexError, FileNotFoundError):
    print("The report folder path was either not given or is not a valid directory. Please try the script again.")
    print("Script usage_report: python /path/csv_merge.py /path/reports [/path/standardize_formats.csv]")
    exit()

# Gets paths for the data files used in this script, which should be in the report folder.
# If any are not found, prints an error and quits the script.
formats_by_aip_report = False
formats_report = False
usage_report = False

for file in os.listdir('.'):
    # This CSV has one line per AIP and unique format. AIPs have multiple rows. Allows aggregating collection and AIP
    # format inforamtion without unpacking lists of ids.
    if file.startswith('archive_formats_by_aip') and file.endswith('.csv'):
        formats_by_aip_report = file
    # This CSV has one line per unique format. Allows aggregating file count information.
    elif file.startswith('archive_formats_') and file.endswith('.csv'):
        formats_report = file
    # This CSV has user and group ingest information.
    elif file.startswith('usage_report_') and file.endswith('.csv'):
        usage_report = file

if not formats_by_aip_report:
    print("Could not find archive formats_by_aip_report csv in the report folder.")
    if not usage_report:
        print("Could not find usage_report report csv in the report folder.")
    exit()

# Increases the size of csv fields to handle long aip lists.
# Gets the maximum size that doesn't give an overflow error.
while True:
    try:
        csv.field_size_limit(sys.maxsize)
        break
    except OverflowError:
        sys.maxsize = int(sys.maxsize / 10)

# Makes dataframes from both format reports.
df = pd.read_csv(formats_report)
df_aip = pd.read_csv(formats_by_aip_report)

# TODO: this replaces dash in all collection ids. Probably ok (not creating duplicates) but more than want to do.
# Updates collection ids to remove dash, since rbrl-### and rbrl### should be treated as one collection. This
# filters for russell collections: df_aip.loc[df_aip['Group'] == 'russell']['Collection'], but gives error if put it on
# left of = and if put on right for df_aip['Collection'] all other collection ids become NaN. russell_collections =
# russell_collections.str.replace('-', '')
df_aip['Collection'] = df_aip['Collection'].str.replace('-', '')

# Makes a spreadsheet to save results to.
# TODO: Making a tab, adding a header, and saving all the results of calling a function is repeating. Own function?
# TODO: can I sort? Bold the first row? Add a chart? What else can I do besides save to a tab?
wb = openpyxl.Workbook()

# Makes the ARCHive overview report (TBS, AIPs, and Collections by group) and saves to the report spreadsheet.

# Gets the data for the overview as a dataframe.
overview = archive_overview()

# Renames the sheet made when starting a workbook to ARCHive Overview.
ws1 = wb.active
ws1.title = "ARCHive Overview"

# Adds a header row to the sheet.
ws1.append(['Group', 'Size (TBs)', 'AIPs', 'Collections', 'Files (inflated)'])

# Converts the overview dataframe to a list of lists, one list per row.
# reset_index() includes the index value (the type) and values.tolist() adds the counts.
# TODO: save directly to Excel from dataframe? See reports_pandas.py
# TODO: are these totals really what I want? Or are they duplicating collections and AIPs in more than one category?
overview_rows = overview.reset_index().values.tolist()

# Adds the overview data to the ARCHive Overview tab in the results spreadsheet.
for overview_row in overview_rows:
    ws1.append(overview_row)


# Makes the format types report and saves to the spreadsheet.
# Count of unique collections, unique AIPs, and (some inflation) files for each format type.

# Creates the subtotals by format type using dataframes and combines them into a single dataframe.
collection_type = df_aip.groupby('Format_Type')['Collection'].nunique()
aip_type = df_aip.groupby('Format_Type')['AIP'].nunique()
file_type = df.groupby('Format_Type')['File_Count'].sum()
type_frames = [collection_type, aip_type, file_type]
types_combined = pd.concat(type_frames, axis=1)

# TODO: add percentages?
# Adds the column totals to the format type dataframes.
types_combined.loc['total'] = [collection_type.sum(), aip_type.sum(), file_type.sum()]

# Converts the format standardized name dataframe to a list of lists, one list per row.
# reset_index() includes the index value (the type) and values.tolist() adds the counts.
# TODO: save directly to Excel from dataframe? See reports_pandas.py
type_rows = types_combined.reset_index().values.tolist()

# Adds the format type data to a tab in the results spreadsheet, with a header row.
ws2 = wb.create_sheet(title="type_counts")
ws2.append(['Format Type', 'Collection Count', 'AIP Count', 'File Count'])
for type_row in type_rows:
    ws2.append(type_row)


# Makes the format standardized name report and saves to the spreadsheet.
# Count of unique collections, unique AIPs, and (some inflation) files for each format standardized name.

# Creates the subtotals by format standardized name using dataframes and combines them into a single dataframe.
collection_name = df_aip.groupby('Format_Standardized_Name')['Collection'].nunique()
aip_name = df_aip.groupby('Format_Standardized_Name')['AIP'].nunique()
file_name = df.groupby('Format_Standardized_Name')['File_Count'].sum()
name_frames = [collection_name, aip_name, file_name]
names_combined = pd.concat(name_frames, axis=1)

# TODO: add percentages?
# Adds the column totals to the format standardized name dataframe.
names_combined.loc['total'] = [collection_name.sum(), aip_name.sum(), file_name.sum()]

# Converts the format standardized name dataframe to a list of lists, one list per row.
# reset_index() includes the index value (the type) and values.tolist() adds the counts.
name_rows = names_combined.reset_index().values.tolist()

# Adds the format standardized name data to a tab in the results spreadsheet, with a header row.
ws3 = wb.create_sheet(title="name_counts")
ws3.append(['Format Standardized Name', 'Collection Count', 'AIP Count', 'File Count'])
for name_row in name_rows:
    ws3.append(name_row)


# Gets the current date, formatted YYYYMM, to use in naming the merged file.
today = datetime.datetime.now().strftime("%Y-%m")

# Can save after each tab if want. Do not save, change the tab, and re-save or it will overwrite.
wb.save(f"ARCHive Format Report_{today}.xlsx")

# # TODO: Reports I was making with pandas that are not included here
# # Are these helpful or would we just go back to the main spreadsheet?
# # Format type then by group
# # Format type then by name
# # Format name then by group
#
# # TODO: ideas for future development
# # A tab with the most common formats, however that is defined.
# # A way to compare one spreadsheet to another to show change since the last report.
# # Adding in size, if Shawn can update the format report.
# # Calculate the number of individual formats in a name or type grouping?
# # Calculate the average amount of format variety in a collection or AIP?
# # Compare to the NARA risk framework?
